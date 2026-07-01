"""
Отчёт по оплатам: фильтрация, статистика, графики, экспорт.
"""
import csv
import io
from datetime import date
from decimal import Decimal

from django.db.models import Q, Sum
from django.utils import timezone

from core.models import Direction, Payment, PaymentType, Student


PAYMENT_TYPE_LABELS = dict(PaymentType.choices)


def _default_period() -> tuple[date, date]:
    today = timezone.localdate()
    return today.replace(day=1), today


def apply_payment_filters(
    qs,
    *,
    date_from: date | None = None,
    date_to: date | None = None,
    direction_id=None,
    payment_type: str | None = None,
    q: str = '',
):
    if not date_from or not date_to:
        date_from, date_to = _default_period()
    qs = qs.filter(payment_date__gte=date_from, payment_date__lte=date_to)
    if direction_id:
        qs = qs.filter(direction_id=direction_id)
    if payment_type:
        qs = qs.filter(payment_type=payment_type)
    q = (q or '').strip()
    if q:
        qs = qs.filter(
            Q(student__name__icontains=q)
            | Q(student__parent__name__icontains=q)
            | Q(student__parent__phone__icontains=q)
            | Q(notes__icontains=q)
        )
    return qs.distinct()


def get_filtered_payments(
    date_from: date | None = None,
    date_to: date | None = None,
    direction_id=None,
    payment_type: str | None = None,
    q: str = '',
):
    qs = Payment.objects.select_related('student', 'direction', 'student__parent')
    return apply_payment_filters(
        qs,
        date_from=date_from,
        date_to=date_to,
        direction_id=direction_id,
        payment_type=payment_type,
        q=q,
    ).order_by('-payment_date', 'student__name')


def get_payments_summary(payments_qs) -> dict:
    total = payments_qs.aggregate(s=Sum('amount'))['s'] or Decimal('0')
    count = payments_qs.count()
    return {
        'total': total,
        'count': count,
        'avg': (total / count) if count else Decimal('0'),
    }


def _chart_bars(items: list[dict], limit: int = 12) -> list[dict]:
    sorted_items = sorted(items, key=lambda x: x['value'], reverse=True)[:limit]
    max_val = max((x['value'] for x in sorted_items), default=1) or 1
    total = sum(x['value'] for x in sorted_items) or 1
    return [
        {
            **item,
            'value_fmt': f"{item['value']:,.0f}".replace(',', ' '),
            'percent': round(item['value'] / total * 100, 1),
            'width': round(item['value'] / max_val * 100, 1),
        }
        for item in sorted_items
    ]


def get_payments_charts(payments_qs) -> dict:
    by_direction_map: dict = {}
    by_type_map: dict[str, Decimal] = {}

    for p in payments_qs:
        did = str(p.direction_id)
        if did not in by_direction_map:
            by_direction_map[did] = {
                'label': p.direction.name,
                'direction_id': p.direction_id,
                'value': 0.0,
            }
        by_direction_map[did]['value'] += float(p.amount)
        type_label = p.get_payment_type_display()
        by_type_map[type_label] = by_type_map.get(type_label, Decimal('0')) + p.amount

    by_direction = _chart_bars(list(by_direction_map.values()))
    by_type = _chart_bars([
        {'label': label, 'value': float(val)}
        for label, val in by_type_map.items()
    ])

    return {
        'by_direction': by_direction,
        'by_type': by_type,
    }


def export_payments_excel(payments_qs) -> bytes:
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Оплаты'

    headers = ['Дата', 'Ученик', 'Направление', 'Тип', 'Сумма', 'Примечание']
    header_fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, p in enumerate(payments_qs, 2):
        ws.cell(row=row_idx, column=1, value=p.payment_date.strftime('%d.%m.%Y'))
        ws.cell(row=row_idx, column=2, value=p.student.name)
        ws.cell(row=row_idx, column=3, value=p.direction.name)
        ws.cell(row=row_idx, column=4, value=p.get_payment_type_display())
        ws.cell(row=row_idx, column=5, value=float(p.amount))
        ws.cell(row=row_idx, column=6, value=p.notes or '')

    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_payments_csv(payments_qs) -> str:
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow(['Дата', 'Ученик', 'Направление', 'Тип', 'Сумма', 'Примечание'])
    for p in payments_qs:
        writer.writerow([
            p.payment_date.strftime('%d.%m.%Y'),
            p.student.name,
            p.direction.name,
            p.get_payment_type_display(),
            f'{p.amount:.2f}'.replace('.', ','),
            p.notes or '',
        ])
    return '\ufeff' + output.getvalue()
